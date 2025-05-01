import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.builtins import styles

# from Data_Driven_Testing.writing_data_into_excel import workbook


def getRowCount(file,SheetName):
    workbook=openpyxl.load_workbook(file)
    Sheet = workbook[SheetName]
    return(Sheet.max_row)

def getColumnCount(file,SheetName):
    workbook=openpyxl.load_workbook(file)
    Sheet = workbook[SheetName]
    return(Sheet.max_column)

def ReadData(file,SheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    Sheet = workbook[SheetName]
    return Sheet.cell(rownum,columnno).value

def WriteData(file,SheetName,rownum,colnumno,data):
    workbook=openpyxl.load_workbook(file)
    Sheet = workbook[SheetName]
    Sheet.cell(rownum,colnumno).value = data
    workbook.save(file)

def fillGreenColor(file,SheetName,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    greenFill = PatternFill(start_color='60b212',end_color='60b212',fill_type='solid')
    sheet.cell(rownum,columnno).fill=greenFill
    workbook.save(file)

def fillRedColor(file,SheetName,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    redFill = PatternFill(start_color='ff0000',end_color='ff0000',fill_type='solid')
    sheet.cell(rownum,columnno).fill=redFill
    workbook.save(file)


